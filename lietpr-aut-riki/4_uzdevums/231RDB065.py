import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import csv 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name=[]
# program read information from people.csv file and put all data in name list.
with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row=line.rstrip().split(",") 
        name.append(row)

workbook = load_workbook("salary.xlsx")
sheet = workbook.active
excel_data = []
for row in sheet.iter_rows(values_only=True):
    excel_data.append(list(row))

salary_total = []

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
for row in name:
    input = f"{row[2]} {row[3]}"
    elem = driver.find_element(By.ID, "input")
    elem.clear()
    elem.send_keys(input)
    elem = driver.find_element(By.ID, "execute").click()
    elem = driver.find_element(By.ID, "output")
    output = elem.get_attribute("value")
    salary = 0
    for id in excel_data:
        if output in id:
            salary = salary + id[1]
    salary_total.append([input, salary])

with open('total_salary.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    field = ["name", "total_salary"]
    
    writer.writerow(field)
    for row in salary_total:
        writer.writerow(row)